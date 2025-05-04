import os
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill, Font, Border, Side
from openpyxl.utils import get_column_letter


def optimize_excel():
    """优化核对版数据Excel文件，合并相同人员的单元格和相关信息"""
    try:
        # 获取脚本所在目录
        script_dir = os.path.dirname(os.path.abspath(__file__))
        data_dir = os.path.join(script_dir, "考勤数据")
        
        # 查找核对版数据文件
        input_file = None
        for file in os.listdir(data_dir):
            if "核对版数据" in file:
                input_file = os.path.join(data_dir, file)
                break
        
        if not input_file:
            print("未找到核对版数据文件")
            return False
        
        # 读取Excel文件
        df = pd.read_excel(input_file)
        
        # 按姓名和刷卡日期排序
        df = df.sort_values(['姓名', '刷卡日期'])
        
        # 格式化刷卡日期列为年月日格式
        df['刷卡日期'] = pd.to_datetime(df['刷卡日期']).dt.strftime('%Y-%m-%d')
        
        # 创建输出文件路径
        output_file = os.path.join(data_dir, "考勤稽核数据核对版.xlsx")
        
        # 保存为新的Excel文件
        df.to_excel(output_file, index=False)
        
        # 使用openpyxl加载工作簿进行格式调整
        wb = load_workbook(output_file)
        ws = wb.active
        
        # 获取姓名列索引
        name_col = None
        for col in range(1, ws.max_column + 1):
            if ws.cell(row=1, column=col).value == "姓名":
                name_col = col
                break
                
        # 合并姓名单元格并居中
        current_name = None
        start_row = 2
        for row in range(2, ws.max_row + 1):
            name = ws.cell(row=row, column=name_col).value
            if name != current_name:
                if current_name and row > start_row:
                    ws.merge_cells(start_row=start_row, start_column=name_col, 
                                 end_row=row-1, end_column=name_col)
                    # 设置合并后的单元格垂直居中
                    ws.cell(row=start_row, column=name_col).alignment = Alignment(
                        vertical='center',
                        horizontal='center'
                    )
                current_name = name
                start_row = row
                
        # 合并最后一批姓名
        if current_name and ws.max_row >= start_row:
            ws.merge_cells(start_row=start_row, start_column=name_col,
                         end_row=ws.max_row, end_column=name_col)
            ws.cell(row=start_row, column=name_col).alignment = Alignment(
                vertical='center',
                horizontal='center'
            )
        
        # 获取列索引
        name_col = None
        desc_col = None
        overtime_col = None
        
        for col in range(1, ws.max_column + 1):
            header = ws.cell(row=1, column=col).value
            if header == "姓名":
                name_col = col
            elif header == "异常描述":
                desc_col = col
            elif header == "加班单时数":
                overtime_col = col
        
        if not all([name_col, desc_col, overtime_col]):
            print("未找到必要的列")
            return False
        
        # 创建一个字典来跟踪每个姓名的行范围
        name_ranges = {}
        current_name = None
        start_row = None
        
        # 首先确保数据按姓名排序
        df = df.sort_values('姓名')
        
        # 从第2行开始（跳过标题行）
        for row in range(2, ws.max_row + 1):
            name = ws.cell(row=row, column=name_col).value
            
            if name != current_name:
                # 如果有前一个名字的范围，保存它
                if current_name and start_row:
                    name_ranges[current_name] = (start_row, row - 1)
                
                # 开始新的名字范围
                current_name = name
                start_row = row
        
        # 添加最后一个名字的范围
        if current_name and start_row:
            name_ranges[current_name] = (start_row, ws.max_row)
        
        # 只合并姓名单元格
        for name, (start, end) in name_ranges.items():
            if start != end:  # 只有多行才需要合并
                ws.merge_cells(start_row=start, start_column=name_col, end_row=end, end_column=name_col)
        
        # 合并异常描述列（当下方单元格为空时）
        if desc_col:
            for row in range(2, ws.max_row):
                current_value = ws.cell(row=row, column=desc_col).value
                next_value = ws.cell(row=row+1, column=desc_col).value
                
                if current_value and not next_value:
                    # 找到下方连续空单元格的范围
                    end_row = row + 1
                    while end_row < ws.max_row and not ws.cell(row=end_row+1, column=desc_col).value:
                        end_row += 1
                    
                    # 合并单元格
                    ws.merge_cells(start_row=row, start_column=desc_col, 
                                 end_row=end_row, end_column=desc_col)
                    # 设置合并后的单元格垂直居中
                    ws.cell(row=row, column=desc_col).alignment = Alignment(
                        vertical='center',
                        horizontal='center'
                    )
        
        # 设置单元格格式
        for row in range(1, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)
                
                # 设置对齐方式
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                
                # 设置标题行格式
                if row == 1:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        
        # 调整列宽
        for col in range(1, ws.max_column + 1):
            column_letter = get_column_letter(col)
            ws.column_dimensions[column_letter].width = 15
        
        # 保存优化后的文件
        wb.save(output_file)
        print(f"优化完成！结果已保存至: {output_file}")
        
        # 生成异常报告
        try:
            # 创建异常报告工作表
            report_sheet = wb.create_sheet("异常报告")
            
            # 分析异常数据
            anomaly_df = df[df['异常描述'].notna()]
            
            # 统计异常类型
            anomaly_stats = anomaly_df['异常描述'].value_counts().reset_index()
            anomaly_stats.columns = ['异常类型', '出现次数']
            
            # 写入异常统计
            report_sheet.append(['异常类型', '出现次数'])
            for _, row in anomaly_stats.iterrows():
                report_sheet.append([row['异常类型'], row['出现次数']])
            
            # 添加AI分析结论
            report_sheet.append([''])
            report_sheet.append(['AI分析结论:'])
            
            # 根据异常数据生成分析结论
            if len(anomaly_df) > 0:
                most_common = anomaly_stats.iloc[0]['异常类型']
                report_sheet.append([f"最常见的异常类型是: {most_common}"])
                
                # 按部门分析异常
                dept_stats = anomaly_df.groupby('部门')['异常描述'].count().sort_values(ascending=False)
                report_sheet.append([''])
                report_sheet.append(['各部门异常数量统计:'])
                for dept, count in dept_stats.items():
                    report_sheet.append([dept, count])
                
                report_sheet.append([''])
                report_sheet.append(["建议: 重点关注异常高发的部门和异常类型"]) 
            else:
                report_sheet.append(["未发现异常记录"]) 
            
            # 保存包含报告的文件
            wb.save(output_file)
            print(f"异常报告已生成并保存至: {output_file}")
            
        except Exception as e:
            print(f"生成异常报告时出错: {str(e)}")
            
        return True
        
    except Exception as e:
        print(f"优化文件时出错: {str(e)}")
        return False


if __name__ == "__main__":
    optimize_excel()
