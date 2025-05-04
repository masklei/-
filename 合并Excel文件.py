import pandas as pd
import os
from openpyxl import Workbook

def get_files_from_attendance_folder():
    """从考勤数据文件夹获取稽查结果文件"""
    script_dir = os.path.dirname(os.path.abspath(__file__))
    attendance_dir = os.path.join(script_dir, "考勤数据")
    
    if not os.path.exists(attendance_dir):
        raise FileNotFoundError("考勤数据文件夹不存在")
    
    files = os.listdir(attendance_dir)
    night_file = next((f for f in files if "夜班稽查结果" in f), None)
    day_file = next((f for f in files if "白班稽查结果" in f), None)
    
    if not night_file or not day_file:
        raise FileNotFoundError("未找到完整的稽查结果文件")
    
    return (
        os.path.join(attendance_dir, night_file),
        os.path.join(attendance_dir, day_file)
    )

def merge_excel_files(file1, file2, output_file):
    """合并两个Excel文件"""
    try:
        # 读取文件
        df1 = pd.read_excel(file1)
        df2 = pd.read_excel(file2)
        
        # 转换刷卡时间格式
        for df in [df1, df2]:
            if '刷卡时间' in df.columns:
                df['刷卡时间'] = pd.to_datetime(df['刷卡时间']).dt.strftime('%H:%M:%S')
        
        # 合并数据
        merged_df = pd.concat([df1, df2], ignore_index=True)
        
        # 确保输出目录存在
        os.makedirs(os.path.dirname(output_file), exist_ok=True)
        
        # 使用openpyxl保存以保留合并单元格
        wb = Workbook()
        ws = wb.active
        
        # 写入表头
        for col_num, column in enumerate(merged_df.columns, 1):
            ws.cell(row=1, column=col_num, value=column)
        
        # 写入数据并处理合并单元格
        current_group = None
        start_row = 2
        
        for row_num, row in merged_df.iterrows():
            group_key = f"{row['姓名']}_{row['刷卡日期']}"
            
            if group_key != current_group:
                if current_group is not None:
                    # 合并上一个组的异常描述单元格
                    ws.merge_cells(start_row=start_row, start_column=merged_df.columns.get_loc('异常描述')+1,
                                  end_row=row_num+1, end_column=merged_df.columns.get_loc('异常描述')+1)
                
                current_group = group_key
                start_row = row_num + 2
            
            # 写入数据
            for col_num, value in enumerate(row, 1):
                ws.cell(row=row_num+2, column=col_num, value=value)
        
        # 保存文件
        wb.save(output_file)
        print(f"文件已成功合并并保存到: {output_file}")
        return True
    except Exception as e:
        print(f"合并文件时出错: {str(e)}")
        return False

if __name__ == "__main__":
    try:
        # 自动获取文件
        night_file, day_file = get_files_from_attendance_folder()
        
        # 设置输出路径
        output_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "考勤数据")
        os.makedirs(output_dir, exist_ok=True)
        output_file = os.path.join(output_dir, "合并结果.xlsx")
        
        # 合并文件
        merge_excel_files(night_file, day_file, output_file)
        
    except Exception as e:
        print(f"程序运行出错：{str(e)}")
