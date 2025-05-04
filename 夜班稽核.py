import pandas as pd
import tkinter as tk
from tkinter import filedialog
import datetime
import re
from collections import defaultdict
import os
import concurrent.futures
from functools import partial


def select_file():
    """选择输入文件"""
    root = tk.Tk()
    root.withdraw()
    file_path = filedialog.askopenfilename(title="选择考勤数据文件",
                                           filetypes=[("Excel files", "*.xlsx"), ("CSV files", "*.csv"),
                                                      ("All files", "*.*")])
    root.destroy()
    return file_path


def is_night_shift(shift):
    """判断是否为夜班"""
    return '夜班' in str(shift)


def parse_datetime(date_str, time_str):
    """解析日期和时间字符串为datetime对象"""
    try:
        # 确保输入不是空值或数字
        if pd.isna(date_str) or pd.isna(time_str) or isinstance(date_str, (int, float)) or isinstance(time_str, (int, float)):
            return None
            
        # 转换为datetime对象
        date_obj = pd.to_datetime(date_str, errors='coerce')
        time_obj = pd.to_datetime(time_str, errors='coerce')
        
        # 检查转换是否成功
        if pd.isna(date_obj) or pd.isna(time_obj):
            return None
            
        return datetime.datetime.combine(date_obj.date(), time_obj.time())
    except Exception as e:
        print(f"解析日期时间错误: {e}, date_str={date_str}, time_str={time_str}")
        return None


def get_time_diff_minutes(time1, time2):
    """计算两个时间之间的分钟差"""
    diff = (time2 - time1).total_seconds() / 60
    return abs(diff)


def parse_time(time_val):
    """解析时间值为datetime.time对象"""
    if pd.isna(time_val):
        return None
        
    if isinstance(time_val, datetime.datetime):
        return time_val.time()
    elif isinstance(time_val, datetime.time):
        return time_val
    elif isinstance(time_val, str):
        try:
            return datetime.datetime.strptime(time_val, '%H:%M:%S').time()
        except ValueError:
            try:
                return datetime.datetime.strptime(time_val, '%H:%M').time()
            except ValueError:
                return None
    return None


def is_time_covered_by_leave(work_time, actual_time, leave_start, leave_end):
    """检查迟到时间是否被请假时间段覆盖"""
    if pd.isna(leave_start) or pd.isna(leave_end):
        return False
    
    # 转换为time对象以便比较
    work_time = parse_time(work_time)
    actual_time = parse_time(actual_time)
    leave_start = parse_time(leave_start)
    leave_end = parse_time(leave_end)
    
    if not work_time or not actual_time or not leave_start or not leave_end:
        return False
    
    # 检查迟到时间段是否被请假时间段覆盖
    # 将time对象转换为分钟数进行比较
    def time_to_minutes(t):
        return t.hour * 60 + t.minute
    
    work_minutes = time_to_minutes(work_time)
    actual_minutes = time_to_minutes(actual_time)
    leave_start_minutes = time_to_minutes(leave_start)
    leave_end_minutes = time_to_minutes(leave_end)
    
    # 如果请假时间覆盖了迟到时间段，则返回True
    return leave_start_minutes <= work_minutes and leave_end_minutes >= actual_minutes


def check_night_shift_anomalies(df):
    """检查夜班考勤异常"""
    # 初始化结果 DataFrame
    result_df = pd.DataFrame()  # 显式初始化为空 DataFrame
    
    # 初始化额外的列
    extra_columns = ['外出时间', '进入时间', '外出时长', '连续进入时间1', '连续进入时间2']  # 删除'实际加班时长'
    for col in extra_columns:
        df[col] = None

    # 按员工分组
    employee_groups = df.groupby('姓名')

    for name, group in employee_groups:
        # 筛选夜班记录
        night_shifts = group[group['班别'].apply(is_night_shift)]
        if night_shifts.empty:
            continue

        # 按日期和时间排序
        night_shifts = night_shifts.sort_values(['刷卡日期', '刷卡时间'])

        # 重新组织夜班记录（以12点为分界）
        shift_records = defaultdict(list)
        for _, row in night_shifts.iterrows():
            dt = parse_datetime(row['刷卡日期'], row['刷卡时间'])
            if not dt:
                continue

            # 忽略12:00-18:00的打卡记录
            if datetime.time(12, 0) <= dt.time() <= datetime.time(18, 0):
                continue

            # 确定记录属于哪个夜班班次（以12点为分界）
            if dt.time() >= datetime.time(12, 0):
                shift_date = dt.date()
            else:
                shift_date = dt.date() - datetime.timedelta(days=1)

            shift_records[shift_date].append({
                'datetime': dt,
                'type': row['来源'],
                'direction': row['刷卡机'],
                'row': row
            })

        # 检查每个夜班班次的异常
        for shift_date, records in shift_records.items():
            # 按时间排序
            records = sorted(records, key=lambda x: x['datetime'])
            
            # 保存原始记录，用于输出完整打卡记录
            original_records = records.copy()
            
            # 获取夜班的上班时间(20:00)和下班时间(04:00)
            work_start_time = datetime.time(20, 0)
            work_end_time = datetime.time(4, 0)

            # 获取加班单信息
            # has_overtime_form = False
            # overtime_form_start_time = None
            # overtime_form_end_time = None
            # overtime_form_hours = 0
            has_overtime_form = False
            overtime_form_start_time = None
            overtime_form_end_time = None
            overtime_form_hours = 0
            
            if records and 'row' in records[0]:
                row = records[0]['row']
                if not pd.isna(row.get('加班单开始时间')) and not pd.isna(row.get('加班单结束时间')):
                    has_overtime_form = True
                    overtime_form_start_time = parse_time(row.get('加班单开始时间'))
                    overtime_form_end_time = parse_time(row.get('加班单结束时间'))
                    overtime_form_hours = float(row.get('加班单时数', 0)) if not pd.isna(row.get('加班单时数', 0)) else 0
            
            # 加班开始时间和结束时间
            # 如果有加班单，使用加班单的时间，否则使用默认时间
            overtime_start_time = datetime.time(4, 40)
            overtime_end_time = datetime.time(8, 10)
            
            # 分类记录：上班前、工作时间内、下班后、加班时间内
            before_work_records = []
            work_time_records = []
            after_work_records = []
            # overtime_records = []  # 删除加班时间记录
            
            for record in records:
                record_time = record['datetime'].time()
                # 上班前记录 (< 20:00)
                if record_time < work_start_time:
                    before_work_records.append(record)
                # 下班后记录 (> 04:00 且 < 05:10)
                elif record_time > work_end_time and record_time < overtime_start_time:
                    after_work_records.append(record)
                # 加班时间记录 (>= 05:10 且 <= 08:10)
                # elif overtime_start_time <= record_time <= overtime_end_time:
                #     overtime_records.append(record)
                # 工作时间记录 (>= 20:00 或 <= 04:00)
                else:
                    work_time_records.append(record)
            
            # 处理上班前记录 - 只保留最后一次进入记录
            last_in_before_work = None
            for record in reversed(before_work_records):
                if record['direction'] == '进':
                    last_in_before_work = record
                    break
            
            # 处理下班后记录 - 只保留第一次打卡记录
            first_after_work = after_work_records[0] if after_work_records else None
            
            # 合并有效记录
            filtered_records = []
            if last_in_before_work:
                filtered_records.append(last_in_before_work)
            filtered_records.extend(work_time_records)
            if first_after_work:
                filtered_records.append(first_after_work)
            # filtered_records.extend(overtime_records)  # 删除加班时间记录
            
            # 重新按时间排序
            filtered_records = sorted(filtered_records, key=lambda x: x['datetime'])
            
            # 处理两分钟内多次打卡的情况
            processed_records = []
            i = 0
            while i < len(filtered_records):
                current = filtered_records[i]
                j = i + 1
                duplicates = [current]

                # 查找2分钟内的所有记录
                while j < len(filtered_records) and (filtered_records[j]['datetime'] - current['datetime']).total_seconds() <= 120:
                    duplicates.append(filtered_records[j])
                    j += 1

                # 获取当前时间
                current_time = current['datetime'].time()
                
                # 判断是否在工作时间内或加班时间内
                is_work_time = (work_start_time <= current_time or current_time <= work_end_time)
                # is_overtime = overtime_start_time <= current_time <= overtime_end_time
                # 不再检查进出方向不一致，只保留最后一条记录
                if (is_work_time):
                    # 只保留最后一条记录
                    processed_records.append(duplicates[-1])
                else:
                    # 非工作时间，保留所有记录
                    for r in duplicates:
                        processed_records.append(r)

                i = j

            # 使用处理后的记录进行异常检测
            records = processed_records

            # 检查其他异常
            anomalies = []
            descriptions = []

            if records:
                # 获取请假信息
                leave_start = None
                leave_end = None
                if 'row' in records[0] and '请假开始时间' in records[0]['row'] and '请假结束时间' in records[0]['row']:
                    leave_start = records[0]['row'].get('请假开始时间')
                    leave_end = records[0]['row'].get('请假结束时间')
                
                # 检查上班打卡 - 以上班前最后一次进入记录为准
                last_in_before_work = None
                for record in records:
                    if record['direction'] == '进' and record['datetime'].time() < work_start_time:
                        last_in_before_work = record
                
                # 如果没有上班前的进入记录，查找最早的进入记录
                first_in = None
                for record in records:
                    if record['direction'] == '进':
                        first_in = record
                        break
                
                # 检查迟到是否被请假覆盖
                late_covered_by_leave = False
                if first_in and leave_start is not None and leave_end is not None:
                    late_covered_by_leave = is_time_covered_by_leave(work_start_time, first_in['datetime'].time(), leave_start, leave_end)
                
                # 判断迟到
                if not last_in_before_work and first_in and first_in['datetime'].time() > datetime.time(20, 1) and not late_covered_by_leave:
                    anomalies.append("首次进入超时")
                    descriptions.append(f"首次进入时间为{first_in['datetime'].strftime('%H:%M:%S')}，超过20:01")

                # 下班判定逻辑
                # 无加班单：以4:00后第一条"刷卡机=出"记录作为下班时间，后续打卡记录忽略
                # 有加班单：以加班单结束时间后的第一条"刷卡机=出"记录作为下班时间，后续打卡记录忽略
                # 只保留无加班单逻辑
                first_out_after_work = None
                for record in records:
                    if record['direction'] == '出' and record['datetime'].time() > work_end_time:
                        first_out_after_work = record
                        break
            # 删除所有与加班时长异常相关的异常判定与描述输出
            # 检查下班打卡
            first_out_after_overtime = None
            
            # 如果没有加班单，以4:00后第一条出记录为下班时间
            if not has_overtime_form:
                for record in records:
                    if record['direction'] == '出' and record['datetime'].time() > work_end_time:
                        first_out_after_work = record
                        break
            # 如果有加班单，以加班单结束时间后的第一条出记录为下班时间
            else:
                for record in records:
                    if record['direction'] == '出' and record['datetime'].time() > overtime_end_time:
                        first_out_after_overtime = record
                        break
            
            # 如果没有下班后的出记录，查找最后一次出记录
            last_out = None
            for record in reversed(records):
                if record['direction'] == '出':
                    last_out = record
                    break
            
            # 添加对提前下班的检查
            if not has_overtime_form and last_out and last_out['datetime'].time() < work_end_time and not first_out_after_work:
                anomalies.append("提前下班")
                descriptions.append(f"最后一次出卡时间为{last_out['datetime'].strftime('%H:%M:%S')}，早于正常下班时间04:00")

            # 检查异常1：工作期间出入时间差大于15分钟 - 只在工作时间和加班时间内判断
            i = 0
            while i < len(records) - 1:
                current_time = records[i]['datetime'].time()
                next_time = records[i+1]['datetime'].time()
                
                # 判断是否在工作时间内或加班时间内
                current_in_work_time = (work_start_time <= current_time or current_time <= work_end_time)
                current_in_overtime = overtime_start_time <= current_time <= overtime_end_time
                next_in_work_time = (work_start_time <= next_time or next_time <= work_end_time)
                next_in_overtime = overtime_start_time <= next_time <= overtime_end_time
                
                if (current_in_work_time or current_in_overtime) and (next_in_work_time or next_in_overtime):
                    if records[i]['direction'] == '出' and records[i + 1]['direction'] == '进':
                        time_diff = get_time_diff_minutes(records[i]['datetime'], records[i + 1]['datetime'])
                        if time_diff > 15:
                            # 检查外出时间是否被请假覆盖
                            out_time_covered_by_leave = False
                            if leave_start is not None and leave_end is not None:
                                out_time_covered_by_leave = is_time_covered_by_leave(records[i]['datetime'].time(), 
                                                                                    records[i + 1]['datetime'].time(), 
                                                                                    leave_start, leave_end)
                            
                            # 记录外出信息，无论是否异常
                            for record in original_records:
                                if 'row' in record:
                                    record['row']['外出时间'] = records[i]['datetime'].strftime('%H:%M:%S')
                                    record['row']['进入时间'] = records[i + 1]['datetime'].strftime('%H:%M:%S')
                                    record['row']['外出时长'] = int(time_diff)
                            
                            # 只有当外出时间未被请假覆盖时才标记为异常
                            if not out_time_covered_by_leave:
                                anomalies.append("工作时间外出超过15分钟")
                                descriptions.append(f"外出时间为{records[i]['datetime'].strftime('%H:%M:%S')}，"
                                                        f"再次进入时间为{records[i + 1]['datetime'].strftime('%H:%M:%S')}，"
                                                        f"外出时长{int(time_diff)}分钟")
                i += 1

            # 检查异常2：有进无出 - 只在工作时间和加班时间内判断
            i = 0
            while i < len(records) - 1:
                current_time = records[i]['datetime'].time()
                next_time = records[i+1]['datetime'].time()
                
                # 判断是否在工作时间内或加班时间内
                current_in_work_time = (work_start_time <= current_time or current_time <= work_end_time)
                current_in_overtime = overtime_start_time <= current_time <= overtime_end_time
                next_in_work_time = (work_start_time <= next_time or next_time <= work_end_time)
                next_in_overtime = overtime_start_time <= next_time <= overtime_end_time
                
                if (current_in_work_time or current_in_overtime) and (next_in_work_time or next_in_overtime):
                    if records[i]['direction'] == '进' and records[i + 1]['direction'] == '进':
                        # 记录连续进入时间
                        for record in original_records:
                            if 'row' in record:
                                record['row']['连续进入时间1'] = records[i]['datetime'].strftime('%H:%M:%S')
                                record['row']['连续进入时间2'] = records[i + 1]['datetime'].strftime('%H:%M:%S')
                        
                        anomalies.append("有进无出")
                        descriptions.append(f"在{records[i]['datetime'].strftime('%H:%M:%S')}进入后，"
                                                f"在{records[i + 1]['datetime'].strftime('%H:%M:%S')}再次进入，无出记录")
                i += 1

            # 检查异常3：有出无进 - 只在工作时间和加班时间内判断
            i = 0
            while i < len(records) - 1:
                current_time = records[i]['datetime'].time()
                next_time = records[i+1]['datetime'].time()
                
                # 判断是否在工作时间内或加班时间内
                current_in_work_time = (work_start_time <= current_time or current_time <= work_end_time)
                current_in_overtime = overtime_start_time <= current_time <= overtime_end_time
                next_in_work_time = (work_start_time <= next_time or next_time <= work_end_time)
                next_in_overtime = overtime_start_time <= next_time <= overtime_end_time
                
                if (current_in_work_time or current_in_overtime) and (next_in_work_time or next_in_overtime):
                    if records[i]['direction'] == '出' and records[i + 1]['direction'] == '出':
                        anomalies.append("有出无进")
                        descriptions.append(f"在{records[i]['datetime'].strftime('%H:%M:%S')}外出后，"
                                                f"在{records[i + 1]['datetime'].strftime('%H:%M:%S')}再次外出，无进入记录")
                i += 1

            # 加班进入判定和加班时长核算
            if has_overtime_form:
                # 检查4:00是否有出记录
                has_out_at_work_end = False
                for record in records:
                    if record['direction'] == '出' and record['datetime'].time() > work_end_time and record['datetime'].time() < overtime_start_time:
                        has_out_at_work_end = True
                        break
                
                # 如果4:00有出记录，则需要在加班开始时间前有进入记录
                if has_out_at_work_end:
                    has_in_before_overtime = False
                    for record in records:
                        if record['direction'] == '进' and record['datetime'].time() < overtime_start_time:
                            has_in_before_overtime = True
                            break
                        
                        if not has_in_before_overtime:
                            anomalies.append("加班未进入")
                            descriptions.append("加班开始前未进入")
                    
                    # 计算实际加班时长
                    # 起始时间：4:40
                    # 结束时间：加班单结束时间后的第一条"刷卡机=出"记录
                    overtime_start_datetime = datetime.datetime.combine(shift_date, overtime_start_time)
                    overtime_end_datetime = None
                    
                    if first_out_after_overtime:
                        overtime_end_datetime = first_out_after_overtime['datetime']
                    elif last_out and last_out['datetime'].time() > overtime_start_time:
                        overtime_end_datetime = last_out['datetime']
                    
                    if overtime_end_datetime:
                        actual_overtime_minutes = get_time_diff_minutes(overtime_start_datetime, overtime_end_datetime)
                        actual_overtime_hours = actual_overtime_minutes / 60
                        
                        # 记录实际加班时长
                        for record in original_records:
                            if 'row' in record:
                                record['row']['实际加班时长'] = round(actual_overtime_hours, 2)
                        
                        # 如果实际加班时长小于加班单时数，标记为异常
                        if actual_overtime_hours < overtime_form_hours:
                            anomalies.append("加班时长不足")
                            descriptions.append(f"实际加班时长{round(actual_overtime_hours, 2)}小时，少于加班单时数{overtime_form_hours}小时")
                else:
                    # 无加班单情况下的加班时长检查
                    overtime_in_records = [r for r in records if r['datetime'].time() >= overtime_start_time
                                        and r['datetime'].time() <= overtime_end_time and r['direction'] == '进']
                    overtime_out_records = [r for r in records if r['datetime'].time() >= overtime_start_time
                                         and r['datetime'].time() <= overtime_end_time and r['direction'] == '出']

                    if overtime_in_records and overtime_out_records:
                        # 计算加班时长 - 取最早的进入和最晚的外出
                        overtime_in = min(overtime_in_records, key=lambda x: x['datetime'])
                        overtime_out = max(overtime_out_records, key=lambda x: x['datetime'])

                        # 如果加班开始时间早于4:40，则按4:40计算
                        start_time = max(overtime_in['datetime'],
                                         datetime.datetime.combine(overtime_in['datetime'].date(), overtime_start_time))

                        overtime_minutes = get_time_diff_minutes(start_time, overtime_out['datetime'])
                        overtime_hours = overtime_minutes / 60
                        
                        # 记录实际加班时长
                        for record in original_records:
                            if 'row' in record:
                                record['row']['实际加班时长'] = round(overtime_hours, 2)
                        
                        if overtime_minutes < 180:  # 3小时 = 180分钟
                            anomalies.append("加班时长不足3小时")
                            descriptions.append(f"加班时长为{int(overtime_minutes)}分钟，不足3小时")

            # 如果有异常，将该班次的所有原始打卡记录添加到结果中
            if anomalies:
                # 获取异常描述
                desc = '；'.join(set(descriptions))
                
                # 将所有原始记录添加到结果中
                for record in original_records:
                    new_row = record['row'].copy()
                    new_row['异常'] = '是'
                    new_row['异常描述'] = desc

                    if result_df.empty:
                        result_df = pd.DataFrame([new_row])
                    else:
                        result_df = pd.concat([result_df, pd.DataFrame([new_row])], ignore_index=True)

    return result_df


def get_matched_file():
    """从考勤数据文件夹获取班别匹配结果文件"""
    script_dir = os.path.dirname(os.path.abspath(__file__))
    attendance_dir = os.path.join(script_dir, "考勤数据")
    
    if not os.path.exists(attendance_dir):
        raise FileNotFoundError("考勤数据文件夹不存在")
    
    files = os.listdir(attendance_dir)
    matched_files = [f for f in files if "班别匹配结果" in f]
    
    if not matched_files:
        raise FileNotFoundError("未找到班别匹配结果文件")
    
    # 取最新的文件
    matched_file = max(matched_files, key=lambda f: os.path.getmtime(os.path.join(attendance_dir, f)))
    return os.path.join(attendance_dir, matched_file)

def process_in_thread(file_path):
    """线程处理函数"""
    try:
        # 读取文件
        if file_path.endswith('.csv'):
            df = pd.read_csv(file_path)
        else:
            df = pd.read_excel(file_path)

        # 处理夜班考勤异常
        result_df = check_night_shift_anomalies(df)

        # 保存结果
        output_path = os.path.join(os.path.dirname(file_path), "夜班稽查结果.xlsx")
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            # 确保输出所有需要的列
            required_columns = ['单位', '部门', '部门CXO-2', '工号', '姓名', '刷卡日期', '刷卡时间', '刷卡机', '班别', 
                              '加班单开始日期', '加班单开始时间', '加班单结束日期', '加班单结束时间', '加班单时数', 
                              '请假开始时间', '请假结束时间', '请假时数', '异常', '异常描述', 
                              '外出时间', '进入时间', '外出时长', '连续进入时间1', '连续进入时间2', '实际加班时长']
            
            # 确保所有列都存在
            for col in required_columns:
                if col not in result_df.columns:
                    result_df[col] = None
            
            # 按照要求的列顺序排列
            result_df = result_df[required_columns]
            
            result_df.to_excel(writer, index=False, sheet_name='夜班异常')
            # 获取工作簿和工作表
            workbook = writer.book
            worksheet = writer.sheets['夜班异常']

            # 合并相同班次的异常描述单元格
            from openpyxl.utils import get_column_letter

            # 获取异常描述列的索引
            desc_col_idx = result_df.columns.get_loc('异常描述') + 1  # Excel列从1开始

            # 按姓名和班次日期分组，合并单元格
            current_name = None
            current_shift_date = None
            start_row = 2  # Excel数据从第2行开始（第1行是表头）

            for i, row in enumerate(result_df.itertuples(), start=2):
                name = getattr(row, '姓名')
                dt = parse_datetime(getattr(row, '刷卡日期'), getattr(row, '刷卡时间'))

                # 确定班次日期（以12点为分界）
                if dt and dt.time() >= datetime.time(12, 0):
                    shift_date = dt.date()
                else:
                    shift_date = dt.date() - datetime.timedelta(days=1) if dt else None

                if name != current_name or shift_date != current_shift_date:
                    # 如果是新的员工或班次，结束上一个合并区域
                    if current_name is not None and i > start_row:
                        cell_range = f"{get_column_letter(desc_col_idx)}{start_row}:{get_column_letter(desc_col_idx)}{i - 1}"
                        worksheet.merge_cells(cell_range)

                    # 开始新的合并区域
                    current_name = name
                    current_shift_date = shift_date
                    start_row = i

            # 处理最后一组
            if current_name is not None and start_row < i:
                cell_range = f"{get_column_letter(desc_col_idx)}{start_row}:{get_column_letter(desc_col_idx)}{i}"
                worksheet.merge_cells(cell_range)

        print(f"处理完成，结果已保存至: {output_path}")
        print(f"共发现 {len(result_df)} 条异常记录")
    except Exception as e:
        print(f"处理文件时出错: {str(e)}")

if __name__ == "__main__":
    try:
        # 自动获取文件
        file_path = get_matched_file()
        
        # 使用线程池处理
        with concurrent.futures.ThreadPoolExecutor() as executor:
            future = executor.submit(process_in_thread, file_path)
            future.result()  # 等待线程完成
            
    except Exception as e:
        print(f"程序运行出错：{str(e)}")
