import pandas as pd
import os
from datetime import datetime, timedelta, time
from openpyxl import load_workbook
from openpyxl.styles import PatternFill


def get_matched_file():
    """从考勤数据文件夹获取班别匹配结果文件"""
    script_dir = os.path.dirname(os.path.abspath(__file__))
    attendance_dir = os.path.join(script_dir, "考勤数据")
    
    if not os.path.exists(attendance_dir):
        raise FileNotFoundError("考勤数据文件夹不存在")
    
    files = os.listdir(attendance_dir)
    matched_files = [f for f in files if "班别匹配结果" in f]
    
    if not matched_files:
        raise FileNotFoundError("未找到班别匹配结果文件")
    
    # 取最新的文件
    matched_file = max(matched_files, key=lambda f: os.path.getmtime(os.path.join(attendance_dir, f)))
    return os.path.join(attendance_dir, matched_file)


def parse_time(time_val):
    """解析时间值为datetime.time对象"""
    if pd.isna(time_val):
        return None
        
    if isinstance(time_val, datetime):
        return time_val.time()
    elif isinstance(time_val, time):
        return time_val
    elif isinstance(time_val, str):
        try:
            return datetime.strptime(time_val, '%H:%M:%S').time()
        except ValueError:
            try:
                return datetime.strptime(time_val, '%H:%M').time()
            except ValueError:
                return None
    return None


def time_diff_minutes(time1, time2):
    """计算两个时间之间的分钟差"""
    if time1 and time2:
        diff = (time2 - time1).total_seconds() / 60
        return abs(diff)
    return 0


def is_time_covered_by_leave(out_time, in_time, leave_start, leave_end):
    """检查外出时间段是否被请假时间段覆盖"""
    if pd.isna(leave_start) or pd.isna(leave_end):
        return False
    
    # 转换为time对象以便比较
    out_time = parse_time(out_time)
    in_time = parse_time(in_time)
    leave_start = parse_time(leave_start)
    leave_end = parse_time(leave_end)
    
    if not out_time or not in_time or not leave_start or not leave_end:
        return False
    
    # 检查外出时间段是否完全被请假时间段覆盖
    # 将time对象转换为分钟数进行比较
    def time_to_minutes(t):
        return t.hour * 60 + t.minute
    
    out_minutes = time_to_minutes(out_time)
    in_minutes = time_to_minutes(in_time)
    leave_start_minutes = time_to_minutes(leave_start)
    leave_end_minutes = time_to_minutes(leave_end)
    
    return leave_start_minutes <= out_minutes and leave_end_minutes >= in_minutes


def process_attendance_data(file_path):
    """处理考勤数据并检测异常"""
    # 读取数据
    try:
        df = pd.read_excel(file_path)
        
        # 检查必要的列是否存在
        required_columns = ['姓名', '刷卡日期', '班别', '刷卡时间', '刷卡机']
        for col in required_columns:
            if col not in df.columns:
                print(f"缺少必要列: {col}")
                return None
        
        # 确保所需列存在
        for col in ['单位', '部门', '部门CXO-2', '工号', '加班单开始日期', '加班单开始时间', 
                   '加班单结束日期', '加班单结束时间', '加班单时数', '请假开始时间', '请假结束时间', '请假时数']:
            if col not in df.columns:
                df[col] = ''
        
        # 转换日期和时间格式
        df['刷卡时间'] = pd.to_datetime(df['刷卡时间'], errors='coerce')
        df['刷卡日期'] = pd.to_datetime(df['刷卡日期'], errors='coerce')
        
        # 提取时间部分
        df['时间'] = df['刷卡时间'].dt.time
        
        # 确保加班和请假时间列的格式正确
        for col in ['加班单开始时间', '加班单结束时间', '请假开始时间', '请假结束时间']:
            if col in df.columns:
                # 先处理NaN值
                df[col] = df[col].apply(lambda x: None if pd.isna(x) else x)
                # 然后转换非NaN值为time对象
                def safe_convert_to_time(x):
                    if x is None:
                        return None
                    try:
                        dt = pd.to_datetime(x, errors='coerce')
                        if pd.isna(dt):
                            return None
                        return dt.time()
                    except:
                        return None
                df[col] = df[col].apply(safe_convert_to_time)
        
    except Exception as e:
        print(f"读取文件出错: {str(e)}")
        return None
    
    # 定义白班的时间界限
    work_start_time = datetime.strptime('08:00', '%H:%M').time()
    work_end_time = datetime.strptime('16:40', '%H:%M').time()
    
    # 按姓名和日期分组处理数据
    result_records = []
    grouped = df.groupby(['姓名', '刷卡日期'])
    
    for (name, date), group in grouped:
        # 筛选白班记录
        if not any('白班' in str(shift) for shift in group['班别'].unique()):
            continue
        
        # 按时间排序
        group = group.sort_values('刷卡时间')
        
        # 初始化异常标记和描述
        has_anomaly = False
        anomaly_desc = []
        
        # 提取该员工当天的所有记录
        records = group.to_dict('records')
        
        # 提取进出记录
        in_records = [r for r in records if '进' in str(r['刷卡机'])]
        out_records = [r for r in records if '出' in str(r['刷卡机'])]
        
        # 1. 上班进入判定（08:00前）
        # 检查是否为休息白班或连班半小时白班
        is_rest_day = any('休息白班' in str(shift) for shift in group['班别'].unique())
        is_half_hour_shift = any('连班半小时白班' in str(shift) for shift in group['班别'].unique())
        
        # 检查是否有恰好8:00的进入记录（针对所有白班类型）
        has_exact_start_time = any(r['时间'] == work_start_time for r in in_records)
        
        before_work_in_records = [r for r in in_records if r['时间'] < work_start_time]
        # 检查迟到是否被请假覆盖
        leave_start = records[0].get('请假开始时间')
        leave_end = records[0].get('请假结束时间')
        late_covered_by_leave = False
        if leave_start is not None and leave_end is not None and not pd.isna(leave_start) and not pd.isna(leave_end):
            # 如果有进入记录，检查是否被请假覆盖
            if in_records:
                late_covered_by_leave = is_time_covered_by_leave(work_start_time, in_records[0]['时间'], leave_start, leave_end)
        
        # 判断迟到情况：
        # 1. 没有8:00前的进入记录
        # 2. 没有恰好8:00的进入记录
        # 3. 迟到没有被请假覆盖
        # 注意：恰好8:00的进入记录已经在has_exact_start_time中判断，不需要重复判断
        if (not before_work_in_records and not has_exact_start_time) and not late_covered_by_leave:
            has_anomaly = True
            anomaly_desc.append("迟到，未在08:00前进入")
        else:
            # 8:00前取最后一条记录作为有效记录
            before_work_in_records = sorted(before_work_in_records, key=lambda r: r['时间'], reverse=True)
            
            # 处理连续进入记录（间隔≤2分钟）
            if len(before_work_in_records) >= 2:
                filtered_records = [before_work_in_records[0]]  # 先取最后一条记录
                for i in range(1, len(before_work_in_records)):
                    current_record = before_work_in_records[i]
                    prev_record = filtered_records[-1]
                    
                    # 计算时间差（分钟）
                    current_time = datetime.combine(date, current_record['时间'])
                    prev_time = datetime.combine(date, prev_record['时间'])
                    time_diff_mins = abs((current_time - prev_time).total_seconds() / 60)
                    
                    # 如果时间差大于2分钟，则保留该记录
                    if time_diff_mins > 2:
                        filtered_records.append(current_record)
                
                before_work_in_records = filtered_records
        
        # 2. 工作时间（08:00~16:40）异常判定
        # 提取工作时间内的记录
        work_time_records = [r for r in records if work_start_time <= r['时间'] <= work_end_time]
        work_time_in_records = [r for r in work_time_records if '进' in str(r['刷卡机'])]
        work_time_out_records = [r for r in work_time_records if '出' in str(r['刷卡机'])]
        
        # 2.1 外出与进入情况
        out_in_pairs = []
        for out_record in work_time_out_records:
            # 找到该出记录后的第一条进记录
            next_in_records = [r for r in work_time_in_records if r['时间'] > out_record['时间']]
            if next_in_records:
                next_in_record = min(next_in_records, key=lambda r: r['时间'])
                out_in_pairs.append((out_record, next_in_record))
            else:
                # 有外出无进入
                # 检查是否为休息白班或连班半小时白班或标准白班且恰好是16:40的出记录（正常下班）
                is_standard_end_time = out_record['时间'] == work_end_time
                is_rest_day_end_time = (is_rest_day or is_half_hour_shift) and is_standard_end_time
                
                # 如果不是正常下班（包括标准白班的16:40出），则进行异常判断
                if not (is_rest_day_end_time or is_standard_end_time):
                    # 检查是否有请假记录覆盖
                    leave_start = records[0].get('请假开始时间')
                    leave_end = records[0].get('请假结束时间')
                    
                    if leave_start is not None and leave_end is not None and not pd.isna(leave_start) and not pd.isna(leave_end):
                        # 有请假记录，检查是否覆盖外出时间
                        if not is_time_covered_by_leave(out_record['时间'], work_end_time, leave_start, leave_end):
                            has_anomaly = True
                            anomaly_desc.append(f"外出未返回且无请假覆盖(外出时间:{out_record['时间']})")
                    else:
                        # 无请假记录
                        has_anomaly = True
                        anomaly_desc.append(f"外出未返回且无请假(外出时间:{out_record['时间']})")
        
        # 处理外出-进入对
        for out_record, in_record in out_in_pairs:
            # 计算外出时长
            out_time = datetime.combine(date, out_record['时间'])
            in_time = datetime.combine(date, in_record['时间'])
            out_duration_minutes = (in_time - out_time).total_seconds() / 60
            
            # 过滤连续打卡（间隔≤2分钟）
            if out_duration_minutes <= 2:
                continue
                
            # 外出时长≤15分钟，忽略
            if out_duration_minutes <= 15:
                continue
                
            # 外出时长>15分钟，检查是否有请假记录覆盖
            leave_start = records[0].get('请假开始时间')
            leave_end = records[0].get('请假结束时间')
            
            if leave_start is None or leave_end is None or pd.isna(leave_start) or pd.isna(leave_end) or not is_time_covered_by_leave(out_record['时间'], in_record['时间'], leave_start, leave_end):
                has_anomaly = True
                anomaly_desc.append(f"外出时长超15分钟(外出时间:{out_record['时间']},进入时间:{in_record['时间']},外出时长:{out_duration_minutes:.0f}分钟)")
                # 添加外出相关信息
                for r in records:
                    r['外出时间'] = out_record['时间']
                    r['进入时间'] = in_record['时间']
                    r['外出时长'] = f"{out_duration_minutes:.0f}分钟"
        
        # 2.2 有进入无外出（工作时间内）
        for in_record in work_time_in_records:
            # 找到该进记录前的最后一条出记录
            prev_out_records = [r for r in work_time_out_records if r['时间'] < in_record['时间']]
            
            # 检查是否有请假记录覆盖
            leave_start = records[0].get('请假开始时间')
            leave_end = records[0].get('请假结束时间')
            leave_covered = False
            if leave_start is not None and leave_end is not None and not pd.isna(leave_start) and not pd.isna(leave_end):
                # 检查请假是否覆盖进入时间
                leave_covered = is_time_covered_by_leave(work_start_time, in_record['时间'], leave_start, leave_end)
            
            # 检查是否是加班结束后的进入记录
            is_after_overtime = False
            if has_overtime and overtime_end_time is not None and not pd.isna(overtime_end_time):
                is_after_overtime = in_record['时间'] >= overtime_end_time
            
            # 只有当同时满足以下条件时才标记为异常：
            # 1. 没有前面的出记录
            # 2. 不是第一条进入记录
            # 3. 没有请假覆盖
            # 4. 不是加班结束后的进入记录
            if not prev_out_records and in_record != work_time_in_records[0] and not leave_covered and not is_after_overtime:
                has_anomaly = True
                anomaly_desc.append(f"有进入无对应外出(进入时间:{in_record['时间']})")
        
        # 2.3 检查全天的进出记录连续性（包括工作时间外）
        # 按时间排序所有进入记录
        sorted_in_records = sorted(in_records, key=lambda r: r['时间'])
        
        # 获取加班信息
        has_overtime = any(r.get('加班单开始时间') is not None and not pd.isna(r.get('加班单开始时间')) for r in records)
        overtime_end_time = next((r.get('加班单结束时间') for r in records if r.get('加班单结束时间') is not None and not pd.isna(r.get('加班单结束时间'))), None)
        
        # 对每个进入记录（除了第一个），检查之前是否有对应的出记录
        for i in range(1, len(sorted_in_records)):
            current_in = sorted_in_records[i]
            prev_in = sorted_in_records[i-1]
            
            # 查找两个进入记录之间是否有出记录
            out_between = [r for r in out_records if prev_in['时间'] < r['时间'] < current_in['时间']]
            
            # 计算两次进入记录之间的时间差（分钟）
            prev_time = datetime.combine(date, prev_in['时间'])
            current_time = datetime.combine(date, current_in['时间'])
            time_diff_mins = (current_time - prev_time).total_seconds() / 60
            
            # 检查是否是加班结束后的进入记录
            is_after_overtime = False
            if has_overtime and overtime_end_time is not None and not pd.isna(overtime_end_time):
                is_after_overtime = prev_in['时间'] >= overtime_end_time
            
            # 检查请假是否覆盖连续进入时间
            leave_start = records[0].get('请假开始时间')
            leave_end = records[0].get('请假结束时间')
            leave_covered = False
            if leave_start is not None and leave_end is not None and not pd.isna(leave_start) and not pd.isna(leave_end):
                # 检查请假是否覆盖连续进入时间
                leave_covered = is_time_covered_by_leave(prev_in['时间'], current_in['时间'], leave_start, leave_end)
            
            # 如果满足以下任一条件，则不标记为异常：
            # 1. 两个连续进入记录之间有出记录
            # 2. 两个连续进入记录都在8:00前（以最后一条为准）
            # 3. 两次进入记录时间间隔不超过2分钟（连续同方向打卡）
            # 4. 前一条进入记录是在加班结束后（不需要有对应的出记录）
            # 5. 连续进入时间被请假覆盖
            if not out_between and not (prev_in['时间'] < work_start_time and current_in['时间'] < work_start_time) and time_diff_mins > 2 and not is_after_overtime and not leave_covered:
                has_anomaly = True
                # 添加更详细的时间信息，包括具体的时间点
                anomaly_desc.append(f"连续进入无中间外出(进入时间:{prev_in['时间']}和{current_in['时间']})")
                
                # 记录连续进入的时间点，即使跨越了工作时间和非工作时间
                
                # 添加外出相关信息到记录中，方便在结果中显示
                for r in records:
                    # 记录连续进入的时间点
                    r['连续进入时间1'] = prev_in['时间']
                    r['连续进入时间2'] = current_in['时间']
        
        # 3. 下班判定（16:40后）
        # 检查是否为休息白班
        is_rest_day = any('休息白班' in str(shift) for shift in group['班别'].unique())
        
        # 检查是否有恰好16:40的出记录（针对休息白班）
        has_exact_end_time = False
        if is_rest_day:
            has_exact_end_time = any(r['时间'] == work_end_time for r in out_records)
        
        # 检查是否有加班单
        has_overtime = any(r.get('加班单开始时间') is not None and not pd.isna(r.get('加班单开始时间')) for r in records)
        
        if has_overtime:
            # 3.1 有加班单
            overtime_start_time = next((r.get('加班单开始时间') for r in records if r.get('加班单开始时间') is not None and not pd.isna(r.get('加班单开始时间'))), None)
            overtime_end_time = next((r.get('加班单结束时间') for r in records if r.get('加班单结束时间') is not None and not pd.isna(r.get('加班单结束时间'))), None)
            overtime_hours = next((float(r.get('加班单时数')) if r.get('加班单时数') and not pd.isna(r.get('加班单时数')) else 0 for r in records), 0)
            
            # 如果是休息白班，全天工作时间视为加班
            if is_rest_day:
                # 计算实际加班时长
                first_in_record = min(in_records, key=lambda r: r['时间']) if in_records else None
                last_out_record = max(out_records, key=lambda r: r['时间']) if out_records else None
                
                if first_in_record and last_out_record:
                    # 检查是否是标准时间打卡（8:00进、16:40出）
                    is_standard_time = (first_in_record['时间'] == work_start_time and 
                                       last_out_record['时间'] == work_end_time)
                    
                    # 如果是标准时间打卡，则不标记为异常
                    if is_standard_time:
                        # 不做任何处理，视为正常
                        pass
                    else:
                        first_in_time = datetime.combine(date, first_in_record['时间'])
                        last_out_time = datetime.combine(date, last_out_record['时间'])
                        
                        # 如果结束时间小于开始时间，说明跨天了
                        if last_out_time < first_in_time:
                            last_out_time += timedelta(days=1)
                        
                        actual_overtime_hours = (last_out_time - first_in_time).total_seconds() / 3600
                        
                        # 只有当实际加班时长小于加班单时数时才添加实际加班时长信息并标记为异常
                        if actual_overtime_hours < float(overtime_hours):
                            for r in records:
                                r['实际加班时长'] = f"{actual_overtime_hours:.2f}小时"
                            has_anomaly = True
                            anomaly_desc.append(f"实际加班时长少于加班单时数(实际:{actual_overtime_hours:.2f}小时,加班单:{overtime_hours}小时)")
                        # 大于等于加班单时数时不填充实际加班时长，也不需要做任何处理
            
            if overtime_start_time is not None and overtime_end_time is not None and not pd.isna(overtime_start_time) and not pd.isna(overtime_end_time):
                # 加班进入判定
                # 检查16:40是否有出记录
                has_out_at_work_end = any(r['时间'] == work_end_time for r in out_records)
                
                if has_out_at_work_end:
                    # 需要在加班开始时间前有进记录
                    # 获取16:40到加班开始时间之间的出记录
                    out_before_overtime = [r for r in out_records if work_end_time < r['时间'] < overtime_start_time]
                    
                    # 如果在这段时间内有外出记录，则需要检查是否有对应的进入记录
                    if out_before_overtime:
                        # 对于每条外出记录，检查是否有后续的进入记录
                        for out_record in out_before_overtime:
                            has_corresponding_in = any(r['时间'] > out_record['时间'] and r['时间'] <= overtime_start_time for r in in_records)
                            if not has_corresponding_in:
                                has_anomaly = True
                                anomaly_desc.append(f"加班前外出未返回(外出时间:{out_record['时间']})")
                    else:
                        # 如果没有外出记录，则检查是否有进入记录
                        has_in_before_overtime = any(r['时间'] <= overtime_start_time for r in in_records if r['时间'] > work_end_time)
                        # 对于休息白班，如果恰好8:00进、16:40出，不需要在加班开始前再次进入
                        is_standard_rest_day_time = False
                        if is_rest_day:
                            first_in_record = min(in_records, key=lambda r: r['时间']) if in_records else None
                            last_out_record = max(out_records, key=lambda r: r['时间']) if out_records else None
                            is_standard_rest_day_time = (first_in_record and last_out_record and 
                                                        first_in_record['时间'] == work_start_time and 
                                                        last_out_record['时间'] == work_end_time)
                        
                        if not has_in_before_overtime and not (is_rest_day and is_standard_rest_day_time):
                            has_anomaly = True
                            anomaly_desc.append("加班开始前未进入")
                
                # 加班时长核算
                # 找到加班结束时间后的第一条出记录
                after_overtime_out_records = [r for r in out_records if r['时间'] >= overtime_end_time]
                
                # 找到16:40后的出记录和进记录
                after_work_out_records = [r for r in out_records if r['时间'] > work_end_time]
                after_work_in_records = [r for r in in_records if r['时间'] > work_end_time]
                
                # 计算实际加班时长
                if after_overtime_out_records:
                    # 16:40和加班结束时间取第一条记录
                    actual_overtime_end = after_overtime_out_records[0]['时间']
                    
                    # 计算实际加班时长
                    overtime_start_dt = datetime.combine(date, overtime_start_time)
                    overtime_end_dt = datetime.combine(date, actual_overtime_end)
                    
                    # 如果结束时间小于开始时间，说明跨天了
                    if overtime_end_dt < overtime_start_dt:
                        overtime_end_dt += timedelta(days=1)
                    
                    actual_overtime_hours = (overtime_end_dt - overtime_start_dt).total_seconds() / 3600
                    
                    # 只有当实际加班时长小于加班单时数时才添加实际加班时长信息
                    if actual_overtime_hours < float(overtime_hours):
                        for r in records:
                            r['实际加班时长'] = f"{actual_overtime_hours:.2f}小时"
                        has_anomaly = True
                        anomaly_desc.append(f"实际加班时长少于加班单时数(实际:{actual_overtime_hours:.2f}小时,加班单:{overtime_hours}小时)")
                else:
                    # 检查是否存在16:40后有出记录但无进记录的情况
                    if after_work_out_records and not after_work_in_records:
                        # 员工在16:40后有出记录但无进记录，且在加班单结束时间后有出记录
                        after_overtime_end_out_records = [r for r in out_records if r['时间'] >= overtime_end_time]
                        if after_overtime_end_out_records:
                            has_anomaly = True
                            anomaly_desc.append("加班期间无进入，有外出")
                            
                            # 计算实际加班时长（从加班开始时间到最后一次出记录）
                            last_out_record = max(out_records, key=lambda r: r['时间'])
                            overtime_start_dt = datetime.combine(date, overtime_start_time)
                            overtime_end_dt = datetime.combine(date, last_out_record['时间'])
                            
                            # 如果结束时间小于开始时间，说明跨天了
                            if overtime_end_dt < overtime_start_dt:
                                overtime_end_dt += timedelta(days=1)
                            
                            actual_overtime_hours = (overtime_end_dt - overtime_start_dt).total_seconds() / 3600
                            
                            # 添加实际加班时长信息
                            for r in records:
                                r['实际加班时长'] = f"{actual_overtime_hours:.2f}小时"
                    # 不再将"加班结束后无出记录"计为异常
        else:
            # 3.2 无加班单
            # 检查是否有16:40前的最后一条出记录（早退判断）
            before_end_out_records = [r for r in out_records if r['时间'] < work_end_time]
            
            # 如果是休息白班或连班半小时白班或标准白班且有恰好16:40的出记录，不判定为早退
            has_exact_end_time_normal = any(r['时间'] == work_end_time for r in out_records)
            if before_end_out_records and not ((is_rest_day or is_half_hour_shift) and has_exact_end_time) and not has_exact_end_time_normal:
                # 检查是否有16:40后的出记录
                after_end_out_records = [r for r in out_records if r['时间'] >= work_end_time]
                if not after_end_out_records:
                    # 检查是否有请假记录覆盖早退时间
                    leave_start = records[0].get('请假开始时间')
                    leave_end = records[0].get('请假结束时间')
                    last_out_time = max(before_end_out_records, key=lambda r: r['时间'])['时间']
                    
                    # 检查早退是否被请假覆盖
                    early_leave_covered = False
                    if leave_start is not None and leave_end is not None and not pd.isna(leave_start) and not pd.isna(leave_end):
                        early_leave_covered = is_time_covered_by_leave(last_out_time, work_end_time, leave_start, leave_end)
                    
                    if not early_leave_covered:
                        has_anomaly = True
                        anomaly_desc.append(f"早退，最后一次出卡时间为{last_out_time}")
            
            # 以16:40后第一条出记录作为下班时间，后续记录忽略
            pass
        
        # 如果有异常，将所有记录添加到结果中
        if has_anomaly:
            # 为每条记录添加异常标记和描述
            for record in records:
                record['异常'] = '是'
                record['异常描述'] = '，'.join(anomaly_desc)
                result_records.append(record)
    
    # 创建结果DataFrame
    if result_records:
        result_df = pd.DataFrame(result_records)
        
        # 确保所有需要的列都存在
        output_columns = ['单位', '部门', '部门CXO-2', '工号', '姓名', '刷卡日期', '刷卡时间', '刷卡机', '班别',
                         '加班单开始日期', '加班单开始时间', '加班单结束日期', '加班单结束时间', '加班单时数',
                         '请假开始时间', '请假结束时间', '请假时数', '异常', '异常描述', '外出时间', '进入时间', '外出时长', 
                         '连续进入时间1', '连续进入时间2', '实际加班时长']
        
        # 确保实际加班时长列始终存在
        if '实际加班时长' not in result_df.columns:
            result_df['实际加班时长'] = ''
        
        # 确保所有需要的列都存在
        for col in output_columns:
            if col not in result_df.columns:
                result_df[col] = ''
        
        # 按指定顺序排列列
        result_df = result_df[output_columns]
        
        # 修改日期和时间格式
        result_df['刷卡日期'] = pd.to_datetime(result_df['刷卡日期']).dt.strftime('%Y-%m-%d')
        result_df['刷卡时间'] = pd.to_datetime(result_df['刷卡时间']).dt.strftime('%H:%M:%S')
        
        return result_df
    else:
        print("未发现异常数据")
        return None


def save_result_to_excel(result_df, output_file):
    """保存结果到Excel文件并格式化"""
    if result_df is None or result_df.empty:
        print("没有异常数据需要保存")
        return
    
    # 保存到Excel
    result_df.to_excel(output_file, index=False)
    
    # 使用openpyxl合并相同员工当天的异常描述单元格
    wb = load_workbook(output_file)
    ws = wb.active
    
    # 获取列名到列索引的映射
    column_indices = {}
    for i, column in enumerate(ws[1]):
        column_indices[column.value] = i + 1  # Excel列从1开始
    
    # 按姓名和日期分组
    groups = {}
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        name_idx = column_indices.get('姓名', 0) - 1
        date_idx = column_indices.get('刷卡日期', 0) - 1
        
        if name_idx >= 0 and date_idx >= 0 and len(row) > max(name_idx, date_idx):
            name = row[name_idx]
            date_str = row[date_idx]
            group_key = f"{name}_{date_str}"
            
            if group_key not in groups:
                groups[group_key] = {'start': i, 'end': i}
            else:
                groups[group_key]['end'] = i
    
    # 合并异常描述单元格
    desc_col_idx = column_indices.get('异常描述')
    if desc_col_idx:
        for group_key, range_info in groups.items():
            start_row = range_info['start']
            end_row = range_info['end']
            
            if start_row != end_row:
                ws.merge_cells(start_row=start_row, start_column=desc_col_idx, 
                              end_row=end_row, end_column=desc_col_idx)
    
    # 保存修改后的Excel文件
    wb.save(output_file)
    print(f"异常数据已保存到: {output_file}")


def main():
    try:
        # 获取班别匹配结果文件
        file_path = get_matched_file()
        print(f"处理文件: {file_path}")
        
        # 处理考勤数据
        result_df = process_attendance_data(file_path)
        
        # 保存结果
        if result_df is not None and not result_df.empty:
            # 确保考勤数据目录存在
            script_dir = os.path.dirname(os.path.abspath(__file__))
            attendance_dir = os.path.join(script_dir, "考勤数据")
            if not os.path.exists(attendance_dir):
                os.makedirs(attendance_dir)
            
            # 保存结果
            output_file = os.path.join(attendance_dir, "白班稽查结果.xlsx")
            save_result_to_excel(result_df, output_file)
        else:
            print("未发现需要保存的异常数据")
            
    except Exception as e:
        print(f"程序运行出错: {str(e)}")


if __name__ == "__main__":
    main()
